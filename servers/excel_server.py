"""MCP server for Excel file operations (openpyxl)."""

import os
from pathlib import Path

from fastmcp import FastMCP

mcp = FastMCP("Excel")


def _resolve_path(path: str) -> Path:
    """Resolve path and enforce workspace boundary."""
    workspace = Path(os.getenv("WORKSPACE_DIR", "./workspace")).resolve()
    target = (workspace / path).resolve()
    try:
        target.relative_to(workspace)
    except ValueError:
        raise PermissionError(f"Access denied: '{path}' is outside workspace")
    return target


@mcp.tool()
def read_excel(path: str, sheet: str | None = None) -> str:
    """Read data from an Excel file.

    Args:
        path: Path to .xlsx file relative to workspace.
        sheet: Sheet name. Default: first sheet.

    Returns:
        Formatted table as text.
    """
    from openpyxl import load_workbook

    target = _resolve_path(path)
    if not target.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = load_workbook(target, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        lines.append("\t".join(str(cell) if cell is not None else "" for cell in row))

    wb.close()
    return "\n".join(lines)


@mcp.tool()
def write_excel(path: str, data: str, sheet: str = "Sheet1") -> str:
    """Write data to an Excel file.

    Args:
        path: Path to .xlsx file relative to workspace.
        data: Tab-separated rows, newline-separated lines.
        sheet: Sheet name. Default: Sheet1.

    Returns:
        Confirmation message.
    """
    from openpyxl import Workbook

    target = _resolve_path(path)

    rows = [line.split("\t") for line in data.strip().split("\n")]

    if target.exists():
        from openpyxl import load_workbook

        wb = load_workbook(target)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.create_sheet(title=sheet)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet

    for row_idx, row_data in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value.strip())

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    wb.close()
    return f"Written {len(rows)} rows to {path}"


@mcp.tool()
def list_sheets(path: str) -> str:
    """List all sheets in an Excel file.

    Args:
        path: Path to .xlsx file relative to workspace.

    Returns:
        List of sheet names.
    """
    from openpyxl import load_workbook

    target = _resolve_path(path)
    if not target.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = load_workbook(target, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return "\n".join(f"- {s}" for s in sheets)


@mcp.tool()
def csv_to_excel(csv_path: str, excel_path: str, sheet: str = "Data") -> str:
    """Convert a CSV file to Excel.

    Args:
        csv_path: Path to .csv file in workspace.
        excel_path: Output .xlsx path.
        sheet: Sheet name. Default: Data.

    Returns:
        Confirmation message.
    """
    import csv

    from openpyxl import Workbook

    csv_target = _resolve_path(csv_path)
    excel_target = _resolve_path(excel_path)

    if not csv_target.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    with open(csv_target, encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    excel_target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_target)
    wb.close()
    return f"Converted {csv_path} → {excel_path}"


if __name__ == "__main__":
    mcp.run()
